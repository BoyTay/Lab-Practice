import tkinter as tk
from tkinter import ttk, messagebox
import re
import os
from openpyxl import load_workbook, Workbook

FILE_PATH = 'excel_dangky.xlsx'

# Tạo file Excel nếu chưa có
if not os.path.exists(FILE_PATH):
    wb = Workbook()
    ws = wb.active
    headers = ["Mã SV", "Họ tên", "Ngày sinh", "Email", "Số điện thoại", "Học kỳ", "Năm học", "Môn học"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = h
    wb.save(FILE_PATH)

# ─── Validation ───────────────────────────────────────────
def validate_masv(P):
    return P.isdigit() or P == ""

def validate_sdt(P):
    return P.isdigit() or P == ""

def validate_hocky(P):
    return P in ("", "1", "2", "3")

def validate_email(email):
    return re.fullmatch(r"[^@]+@[^@]+\.[^@]+", email) is not None

def validate_ngaysinh(ns):
    return re.fullmatch(r"\d{2}/\d{2}/\d{4}", ns) is not None

# ─── Đăng ký ──────────────────────────────────────────────
def dang_ky():
    masv     = entry_masv.get().strip()
    hoten    = entry_hoten.get().strip()
    ngaysinh = entry_ngaysinh.get().strip()
    email    = entry_email.get().strip()
    sdt      = entry_sdt.get().strip()
    hocky    = entry_hocky.get().strip()
    namhoc   = combo_namhoc.get().strip()

    mon_list = []
    if var_python.get(): mon_list.append("Lập trình Python")
    if var_java.get():   mon_list.append("Lập trình Java")
    if var_cnpm.get():   mon_list.append("Công nghệ phần mềm")
    if var_web.get():    mon_list.append("Phát triển ứng dụng web")

    errors = []

    if not masv:
        errors.append("• Chưa nhập Mã số sinh viên")
    elif len(masv) != 7:
        errors.append("• Mã số sinh viên phải đủ 7 số")

    if not hoten:
        errors.append("• Chưa nhập Họ tên")

    if not ngaysinh:
        errors.append("• Chưa nhập Ngày sinh")
    elif not validate_ngaysinh(ngaysinh):
        errors.append("• Ngày sinh phải có định dạng dd/mm/yyyy")

    if not email:
        errors.append("• Chưa nhập Email")
    elif not validate_email(email):
        errors.append("• Email không hợp lệ")

    if not sdt:
        errors.append("• Chưa nhập Số điện thoại")
    elif len(sdt) != 10:
        errors.append("• Số điện thoại phải đủ 10 số")

    if not hocky:
        errors.append("• Chưa nhập Học kỳ")

    if not namhoc:
        errors.append("• Chưa chọn Năm học")

    if not mon_list:
        errors.append("• Chưa chọn môn học nào")

    if errors:
        messagebox.showwarning("Thông báo", "\n".join(errors))
        return

    wb = load_workbook(FILE_PATH)
    ws = wb.active
    for mon in mon_list:
        row = ws.max_row + 1
        data = [masv, hoten, ngaysinh, email, sdt, hocky, namhoc, mon]
        for col, val in enumerate(data, 1):
            ws.cell(row=row, column=col).value = val
    wb.save(FILE_PATH)
    messagebox.showinfo("Thành công", f"Đã đăng ký {len(mon_list)} môn học thành công!")
    clear_fields()

def clear_fields():
    for e in [entry_masv, entry_hoten, entry_ngaysinh, entry_email, entry_sdt, entry_hocky]:
        e.delete(0, tk.END)
    combo_namhoc.set("")
    var_python.set(False)
    var_java.set(False)
    var_cnpm.set(False)
    var_web.set(False)

# ─── Giao diện ────────────────────────────────────────────
root = tk.Tk()
root.title("Đăng ký học phần")
root.geometry("480x420")
root.resizable(False, False)
root.config(bg="#c8f0c8")

vcmd_masv  = root.register(validate_masv)
vcmd_sdt   = root.register(validate_sdt)
vcmd_hocky = root.register(validate_hocky)

# Tiêu đề
tk.Label(root, text="THÔNG TIN ĐĂNG KÝ HỌC PHẦN",
         font=("Arial", 13, "bold"), fg="red", bg="#c8f0c8").grid(
         row=0, column=0, columnspan=3, pady=(10, 8))

labels = ["Mã số sinh viên", "Họ tên", "Ngày sinh", "Email", "Số điện thoại", "Học kỳ", "Năm học"]
for i, lbl in enumerate(labels):
    tk.Label(root, text=lbl, bg="#c8f0c8", anchor="w", width=16).grid(
        row=i+1, column=0, padx=(15, 5), pady=3, sticky="w")

entry_masv    = tk.Entry(root, validate="key", validatecommand=(vcmd_masv,  "%P"), width=30, bg="#e0ffe0")
entry_hoten   = tk.Entry(root, width=30, bg="#e0ffe0")
entry_ngaysinh= tk.Entry(root, width=30, bg="#e0ffe0")
entry_email   = tk.Entry(root, width=30, bg="#e0ffe0")
entry_sdt     = tk.Entry(root, validate="key", validatecommand=(vcmd_sdt,   "%P"), width=30, bg="#e0ffe0")
entry_hocky   = tk.Entry(root, validate="key", validatecommand=(vcmd_hocky, "%P"), width=30, bg="#e0ffe0")
combo_namhoc  = ttk.Combobox(root, values=["2022-2023","2023-2024","2024-2025"], width=28, state="readonly")

entries_list = [entry_masv, entry_hoten, entry_ngaysinh, entry_email, entry_sdt, entry_hocky, combo_namhoc]
for i, widget in enumerate(entries_list):
    widget.grid(row=i+1, column=1, columnspan=2, padx=(0,15), pady=3, sticky="w")

# Placeholder ngày sinh
entry_ngaysinh.insert(0, "dd/mm/yyyy")
entry_ngaysinh.config(fg="gray")
def on_focus_in(e):
    if entry_ngaysinh.get() == "dd/mm/yyyy":
        entry_ngaysinh.delete(0, tk.END)
        entry_ngaysinh.config(fg="black")
def on_focus_out(e):
    if not entry_ngaysinh.get():
        entry_ngaysinh.insert(0, "dd/mm/yyyy")
        entry_ngaysinh.config(fg="gray")
entry_ngaysinh.bind("<FocusIn>",  on_focus_in)
entry_ngaysinh.bind("<FocusOut>", on_focus_out)

# Chọn môn học
tk.Label(root, text="Chọn môn học", bg="#c8f0c8", anchor="w", width=16).grid(
    row=8, column=0, padx=(15,5), pady=3, sticky="nw")

var_python = tk.BooleanVar()
var_java   = tk.BooleanVar()
var_cnpm   = tk.BooleanVar()
var_web    = tk.BooleanVar()

frame_mon = tk.Frame(root, bg="#c8f0c8")
frame_mon.grid(row=8, column=1, columnspan=2, sticky="w")
tk.Checkbutton(frame_mon, text="Lập trình Python",       variable=var_python, bg="#c8f0c8").grid(row=0, column=0, sticky="w", padx=5)
tk.Checkbutton(frame_mon, text="Lập trình Java",         variable=var_java,   bg="#c8f0c8").grid(row=0, column=1, sticky="w", padx=5)
tk.Checkbutton(frame_mon, text="Công nghệ phần mềm",     variable=var_cnpm,   bg="#c8f0c8").grid(row=1, column=0, sticky="w", padx=5)
tk.Checkbutton(frame_mon, text="Phát triển ứng dụng web",variable=var_web,    bg="#c8f0c8").grid(row=1, column=1, sticky="w", padx=5)

# Nút
frame_btn = tk.Frame(root, bg="#c8f0c8")
frame_btn.grid(row=10, column=0, columnspan=3, pady=12)
tk.Button(frame_btn, text="Đăng ký", bg="#00aa00", fg="white", width=10,
          font=("Arial",10,"bold"), command=dang_ky).pack(side="left", padx=20)
tk.Button(frame_btn, text="Thoát",   bg="#cc0000", fg="white", width=10,
          font=("Arial",10,"bold"), command=root.quit).pack(side="left", padx=20)

root.mainloop()